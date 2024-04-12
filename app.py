import openpyxl
import openpyxl.workbook

# Criando a planilha (book)
book = openpyxl.Workbook()

# Visualizando paginas existentes
print(book.sheetnames)

# Criando uma página
book.create_sheet('Frutas')

# Selecionando uma página
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maçã', '2', 'R$4,60'])
frutas_page.append(['Laranja', '7', 'R$2,50'])
frutas_page.append(['Melancia', '5', 'R$12,90'])

#Salvando a planilha
book.save('Planilha de Compras.xlsx')